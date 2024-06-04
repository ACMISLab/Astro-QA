#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/5/4 12:01
# @Author : 桐
# @QQ:1041264242
# 注意事项：

from http import HTTPStatus
import dashscope
import json
import openpyxl
from openpyxl import load_workbook
import re

def call_with_prompt(prompt,m):
    #qwen_max
    if m == 'xxxx':
        response = dashscope.Generation.call(
            model=m,
            # model='qwen1.5-7b-chat',
            prompt=prompt
        )
    else:
        response = dashscope.Generation.call(
            # model=dashscope.Generation.Models.qwen_max,
            model=m,
            prompt=prompt
        )
    # The response status_code is HTTPStatus.OK indicate success,
    # otherwise indicate request is failed, you can get error code
    # and message from code and message.
    if response.status_code == HTTPStatus.OK:
        # print(response.output)  # The output text
        # print(response.usage)  # The usage information
        return response.output['text']
    else:
        print(response.code)  # The error code.
        print(response.message)  # The error message.

if __name__ == '__main__':
    for m in model:
        # 加载Excel文件
        workbook = load_workbook('matching.xlsx')
        sheet = workbook.active

        # 获取所有合并的单元格范围
        merged_ranges = sheet.merged_cells.ranges
        num = 0
        # 检查二维列表是否为完全矩形
        results = []
        # 遍历合并的单元格范围
        for merged_range in merged_ranges:
            group = []
            numbers = [int(match) for match in re.findall(r'\d+', str(merged_range))]
            for index in range(numbers[0], numbers[1] + 1):  # 遍历组内行
                temp = []
                for sub_index, column in enumerate(sheet[index]):  # 第2行即为Excel中的A2行，因为openpyxl的行和列索引是从1开始的
                    if sub_index == 0:
                        continue
                    elif (column.value is None) or (column.value == '') or (sub_index == 5):
                        break
                    else:
                        temp.append(column.value.strip())
                # print(temp)
                group.append(temp)
            # print(group)
            # 获取行数（外部列表的长度）
            num_rows = len(group)
            # 由于是完全矩形，可以获取任意一行的长度作为列数
            num_cols = len(group[0]) if group else 0

            promopt_1 = ''
            promopt_2_temp = ''
            for col in range(num_cols):
                group_str = f'组别{col + 1}术语：'
                if col == 0:
                    promopt_2_temp += f"'这里填写来自组别{col + 1}的术语'"
                else:
                    promopt_2_temp += f",'这里填写来自组别{col + 1}的术语'"

                for row in range(num_rows):
                    if row == 0:
                        group_str = group_str + f"'{group[row][col]}'"
                    else:
                        group_str = group_str + f"、'{group[row][col]}'"
                promopt_1 += group_str + '\n'
            # print(promopt_1)
            promopt_2 = ''
            for row in range(num_rows):
                promopt_2 += f'匹配组{row + 1}：({promopt_2_temp})\n'
            # print(promopt_2)

            promopt = f'''请从以下{num_cols}组天文学术语进行全面考察找出它们之间最合适的匹配关系。注意，同一组内的术语不能相互匹配，而是需要与其他组中的术语进行匹配。每个输出结果都应该是一个包含{num_cols}个术语的元组，这{num_cols}个术语必须分别来自不同的组别。并且，每个术语在整个匹配过程中只能使用一次，以确保形成唯一且最大的连线。\n{promopt_1}输出格式应严格遵循以下样式，并不要有任何多余输出：\n{promopt_2}'''
            # print(promopt)
            # 对每个问题循环5次
            answers_for_question = []
            for id in range(5):
                if id == 0:
                    print(promopt)
                response=call_with_prompt(promopt,m)
                print(f'结果{id}：{response}')
                # 将答案添加到当前问题的答案列表中
                answers_for_question.append(response)
                # 将当前问题的所有答案添加到总结果列表中
            results.append(answers_for_question)
            time.sleep(20)
            # 将结果写入到第10到15列
        for i, answers in enumerate(results, start=2):  # start=2是因为我们从第二行开始写入数据
            for j, answer in enumerate(answers, start=10):  # start=10是因为我们要从第10列开始写入
                sheet.cell(row=i, column=j, value=answer)
                # 保存工作簿
        workbook.save(r'result.xlsx')