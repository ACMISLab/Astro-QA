#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/4/26 11:51
# @Author : 桐
# @QQ:1041264242
# 注意事项：
import ast

import jieba
from nltk.translate.bleu_score import sentence_bleu
from rouge import Rouge
from sacrebleu import corpus_chrf
import openpyxl

restult=[]
result_json={}
result_com_json={}

def Score_beita(bleu_score_avg,rouge_1_score_avg,rouge_2_score_avg,rouge_l_score_avg,chrf_score_avg):
    return 0.3*bleu_score_avg*100+0.1*rouge_1_score_avg*100+0.15*rouge_2_score_avg*100+0.1*rouge_l_score_avg*100+0.35*chrf_score_avg

def DGscore(score_4,score_5,score_multi,score_judge,score_connect,score_terminology,score_shortanswer):
    # print(
    #     f'4选1得分：{score_4 * 1:.2f}\n5选1得分：{score_5 * 1.2:.2f}\n多选得分：{score_multi :.2f}\n判断得分：{score_judge * 0.5:.2f}\n连线得分：{(score_connect * 355 * 3):.2f}\n术语得分：{score_terminology * 5:.2f}\n简答得分：{score_shortanswer * 5:.2f}\n')
    # print(
    #     f'综合得分:{(score_4 * 1 + score_5 * 1.2 + score_multi + score_judge * 0.5 + score_connect * 355 * 3 + score_terminology * 5 + score_shortanswer * 5):.2f}')
    result_com_json["Single_4Choice_com"]=f'{score_4 * 1:.2f}'
    result_com_json["Single_5Choice_com"] = f'{score_5 * 1.2:.2f}'
    result_com_json["Multichoices_com"] = f'{score_multi :.2f}'
    result_com_json["Judge_com"] = f'{score_judge * 0.5:.2f}'
    result_com_json["Connect_com"] = f'{(score_connect * 355 * 3):.2f}'
    result_com_json["Terminology_com"] = f'{score_terminology * 5:.2f}'
    result_com_json["Shortanswer_com"] = f'{score_shortanswer * 5:.2f}'
    result_com_json["com"] = f'{(score_4 * 1 + score_5 * 1.2 + score_multi + score_judge * 0.5 + score_connect * 355 * 3 + score_terminology * 5 + score_shortanswer * 5):.2f}'

def calculate_score(str1, str2):
    # 检查str2中的字母是否都出现在str1中
    for char in str2:
        if char not in str1:
            return 0

            # 检查是否完全匹配
    if sorted(str1) == sorted(str2):
        return 2

        # 计算匹配的字母数量
    match_count = sum(1 for char in str2 if char in str1)
    return match_count * 0.5

def search(answer):
    # 查找第一行中出现的ABCD字母，并添加到str1中
    str_choice = ''
    for char in answer:
        if char in 'ABCD':
            str_choice += char
    return str_choice

def one_score(reference,answer):
    str1=reference  #标准答案
    str2=answer     #生成答案
    # 参考答案和大模型生成答案

    # 使用jieba进行分词
    reference_tokens = " ".join(jieba.cut(str1))
    translation_tokens = " ".join(jieba.cut(str2))
    # print(reference_tokens)
    # print(translation_tokens)
    # 计算BLEU分数
    bleu_score = sentence_bleu([reference_tokens.split()], translation_tokens.split(), weights=(1, 0, 0, 0))
    # print(f"BLEU Score: {bleu_score:.2f}")

    # 对于ROUGE，你可以直接使用整个字符串（不需要分词）
    rouge = Rouge()
    scores = rouge.get_scores([reference_tokens], [translation_tokens], avg=True)
    # print(f"ROUGE-1: {scores['rouge-1']['f']:.2f}")
    # print(f"ROUGE-2: {scores['rouge-2']['f']:.2f}")
    # print(f"ROUGE-L: {scores['rouge-l']['f']:.2f}")

    # 将字符串转换为列表，因为corpus_chrf期望一个参考列表和一个假设列表
    references = [[str1]]
    hypotheses = [str2]
    # 计算CHRF得分
    chrf_score = corpus_chrf(hypotheses, references)
    # print(f'CHRF得分: {chrf_score.score:.2f}')
    return bleu_score,scores['rouge-1']['f'],scores['rouge-2']['f'],scores['rouge-l']['f'],chrf_score.score

def calculate_4choices(min_row,max_row,answer_sheet,test_sheet,weight_sheet):
    """
    计算单4选题的top1、top5的平均acc值。
    """
    sum=0
    temp_num = [0, 0, 0, 0, 0, 0]
    score=[0,0,0,0,0]
    for index,row in enumerate(answer_sheet.iter_rows(min_row,max_row)):  # 假设第一行是标题行，从第二行开始读取
        sum+=1
        # 读取每行中特定列的值
        answer_true = row[0].value.strip()
        weight = weight_sheet[index + min_row][0].value
        flag=False
        for sub_index in range(5):
            try:
                test=test_sheet[index + min_row][sub_index].value.strip()
            except:
                test=''

            if test == answer_true:
                flag=True
                temp_num[sub_index]+=1
                score[sub_index]+=weight
        if True==flag:
            temp_num[5]+=1

    top1_avg_acc=(temp_num[0]+temp_num[1]+temp_num[2]+temp_num[3]+temp_num[4])/sum/5
    top5_avg_acc=temp_num[5]/sum

    weight_score_avg=(score[0]+score[1]+score[2]+score[3]+score[4])/5
    # print(score[0])
    # print(f'四选一题目(acc)：top1：{top1_avg_acc} top5：{top5_avg_acc}')
    # print(f'4选权重得分：{weight_score_avg}')

    result_json["Single_4Choice_Top1"]=top1_avg_acc
    result_json["Single_4Choice_Top5"] = top5_avg_acc

    return weight_score_avg

def calculate_5choices(min_row,max_row,answer_sheet,test_sheet,weight_sheet):
    """
    计算单5选题的top1、top5的平均acc值。
    """
    sum=0
    temp_num = [0, 0, 0, 0, 0, 0]
    score = [0, 0, 0, 0, 0]
    for index,row in enumerate(answer_sheet.iter_rows(min_row,max_row)):  # 假设第一行是标题行，从第二行开始读取
        sum+=1
        # 读取每行中特定列的值
        answer_true = row[1].value.strip()
        weight = weight_sheet[index + min_row][1].value
        flag=False
        for sub_index in range(5):
            try:
                test=test_sheet[index + min_row][5+sub_index].value.strip()
            except:
                test=''

            if test == answer_true:
                flag=True
                temp_num[sub_index]+=1
                score[sub_index] += weight
        if True==flag:
            temp_num[5]+=1

    top1_avg_acc=(temp_num[0]+temp_num[1]+temp_num[2]+temp_num[3]+temp_num[4])/sum/5
    top5_avg_acc=temp_num[5]/sum
    weight_score_avg = (score[0] + score[1] + score[2] + score[3] + score[4]) / 5
    # print(f'五选一题目(acc)：top1：{top1_avg_acc} top5:{top5_avg_acc}')
    # print(f'5选权重得分：{weight_score_avg}')

    result_json["Single_5Choice_Top1"]=top1_avg_acc
    result_json["Single_5Choice_Top5"] = top5_avg_acc

    return weight_score_avg

def calculate_judgechoices(min_row,max_row,answer_sheet,test_sheet,weight_sheet):
    """
    计算判断题的top1、top5的平均acc值。
    """
    sum=0
    temp_num = [0, 0, 0, 0, 0, 0]
    score = [0, 0, 0, 0, 0]
    for index,row in enumerate(answer_sheet.iter_rows(min_row,max_row)):  # 假设第一行是标题行，从第二行开始读取
        sum+=1
        # 读取每行中特定列的值
        answer_true = row[2].value.strip()
        weight = weight_sheet[index + min_row][2].value
        flag=False
        for sub_index in range(5):
            try:
                test=test_sheet[index + min_row][10+sub_index].value.strip()
            except:
                test=''

            if test == answer_true:
                flag=True
                temp_num[sub_index]+=1
                score[sub_index] += weight
        if True==flag:
            temp_num[5]+=1

    top1_avg_acc=(temp_num[0]+temp_num[1]+temp_num[2]+temp_num[3]+temp_num[4])/sum/5
    top5_avg_acc=temp_num[5]/sum
    weight_score_avg = (score[0] + score[1] + score[2] + score[3] + score[4]) / 5
    # print(f'判断题(acc)：top1：{top1_avg_acc} top5:{top5_avg_acc}')
    # print(f'判断题权重得分：{weight_score_avg}')

    result_json["Judge_Top1"]=top1_avg_acc
    result_json["Judge_Top5"] = top5_avg_acc

    return weight_score_avg

def calculate_multichoices(min_row,max_row,answer_sheet,test_sheet):
    """
    计算单多选题的得分。
    """
    temp_score = [0, 0, 0, 0, 0]
    # 按行遍历所有行
    for index,row in enumerate(answer_sheet.iter_rows(min_row,max_row)):  # 假设第一行是标题行，从第二行开始读取
        answer_true = row[3].value.strip()

        # 读取每行中特定列的值
        answer1 = test_sheet[index + min_row][15].value.strip()
        answer2 = test_sheet[index + min_row][16].value.strip()
        answer3 = test_sheet[index + min_row][17].value.strip()
        answer4 = test_sheet[index + min_row][18].value.strip()
        answer5 = test_sheet[index + min_row][19].value.strip()

        choice1 = search(answer1)
        choice2 = search(answer2)
        choice3 = search(answer3)
        choice4 = search(answer4)
        choice5 = search(answer5)

        score1 = calculate_score(answer_true, choice1)
        score2 = calculate_score(answer_true, choice2)
        score3 = calculate_score(answer_true, choice3)
        score4 = calculate_score(answer_true, choice4)
        score5 = calculate_score(answer_true, choice5)

        temp_score[0] += score1
        temp_score[1] += score2
        temp_score[2] += score3
        temp_score[3] += score4
        temp_score[4] += score5

    avg_score=(temp_score[0]+temp_score[1]+temp_score[2]+temp_score[3]+temp_score[4])/5
    # print(f'多选题得分(score):{avg_score}')

    result_json["Multichoices"] = avg_score

    return avg_score

def calculate_terminology(min_row,max_row,answer_sheet,test_sheet):
    """
    计算单术语题的分数
    """
    # 按行遍历所有行
    num = 0
    bleu_score = [0, 0, 0, 0, 0]
    rouge_1_score = [0, 0, 0, 0, 0]
    rouge_2_score = [0, 0, 0, 0, 0]
    rouge_l_score = [0, 0, 0, 0, 0]
    chrf_score = [0, 0, 0, 0, 0]

    for index, row in enumerate(answer_sheet.iter_rows(min_row,max_row)):  # 假设第一行是标题行，从第二行开始读取
        num += 1
        # 读取每行中特定列的值
        reference = row[4].value.strip()  # 假设问题在第六列
        answer1 = test_sheet[index + min_row][20].value.strip()
        answer2 = test_sheet[index + min_row][21].value.strip()
        answer3 = test_sheet[index + min_row][22].value.strip()
        answer4 = test_sheet[index + min_row][23].value.strip()
        answer5 = test_sheet[index + min_row][24].value.strip()

        a1, a2, a3, a4, a5 = one_score(reference, answer1)
        b1, b2, b3, b4, b5 = one_score(reference, answer2)
        c1, c2, c3, c4, c5 = one_score(reference, answer3)
        d1, d2, d3, d4, d5 = one_score(reference, answer4)
        e1, e2, e3, e4, e5 = one_score(reference, answer5)

        bleu_score[0] += a1
        bleu_score[1] += b1
        bleu_score[2] += c1
        bleu_score[3] += d1
        bleu_score[4] += e1

        rouge_1_score[0] += a2
        rouge_1_score[1] += b2
        rouge_1_score[2] += c2
        rouge_1_score[3] += d2
        rouge_1_score[4] += e2

        rouge_2_score[0] += a3
        rouge_2_score[1] += b3
        rouge_2_score[2] += c3
        rouge_2_score[3] += d3
        rouge_2_score[4] += e3

        rouge_l_score[0] += a4
        rouge_l_score[1] += b4
        rouge_l_score[2] += c4
        rouge_l_score[3] += d4
        rouge_l_score[4] += e4

        chrf_score[0] += a5
        chrf_score[1] += b5
        chrf_score[2] += c5
        chrf_score[3] += d5
        chrf_score[4] += e5
        # print(reference)

    # print(num)
    # print(f'bleu_score_avg:{bleu_score[0] / num}  {bleu_score[1] / num}  {bleu_score[2] / num}  {bleu_score[3] / num}  {bleu_score[4] / num}')
    # print(f'rouge_1_score_avg:{rouge_1_score[0] / num}  {rouge_1_score[1] / num}  {rouge_1_score[2] / num}  {rouge_1_score[3] / num}  {rouge_1_score[4] / num}')
    # print(f'rouge_2_score_avg:{rouge_2_score[0] / num}  {rouge_2_score[1] / num}  {rouge_2_score[2] / num}  {rouge_2_score[3] / num}  {rouge_2_score[4] / num}')
    # print(f'rouge_l_score_avg:{rouge_l_score[0] / num}  {rouge_l_score[1] / num}  {rouge_l_score[2] / num}  {rouge_l_score[3] / num}  {rouge_l_score[4] / num}')
    # print(f'chrf_score_avg:{chrf_score[0] / num}  {chrf_score[1] / num}  {chrf_score[2] / num}  {chrf_score[3] / num}  {chrf_score[4] / num}')
    # print("术语解释题：")
    bleu_score_avg=(bleu_score[0]+bleu_score[1]+bleu_score[2]+bleu_score[3]+bleu_score[4]) /num/ 5
    rouge_1_score_avg=(rouge_1_score[0]+rouge_1_score[1]+rouge_1_score[2]+rouge_1_score[3]+rouge_1_score[4])/num/5
    rouge_2_score_avg=(rouge_2_score[0]+rouge_2_score[1]+rouge_2_score[2]+rouge_2_score[3]+rouge_2_score[4])/num/5
    rouge_l_score_avg=(rouge_l_score[0]+rouge_l_score[1]+rouge_l_score[2]+rouge_l_score[3]+rouge_l_score[4])/num/5
    chrf_score_avg=(chrf_score[0]+chrf_score[1]+chrf_score[2]+chrf_score[3]+chrf_score[4])/num/5
    # print(f'bleu_score_avg:{bleu_score_avg}')
    # print(f'rouge_1_score_avg:{rouge_1_score_avg}')
    # print(f'rouge_2_score_avg:{rouge_2_score_avg}')
    # print(f'rouge_l_score_avg:{rouge_l_score_avg}')
    # print(f'chrf_score_avg:{chrf_score_avg}')

    result_json["Terminology_bleu_score_avg"] = bleu_score_avg
    result_json["Terminology_rouge_1_score_avg"] = rouge_1_score_avg
    result_json["Terminology_rouge_2_score_avg"] = rouge_2_score_avg
    result_json["Terminology_rouge_l_score_avg"] = rouge_l_score_avg
    result_json["Terminology_chrf_score_avg"] = chrf_score_avg

    return bleu_score_avg,rouge_1_score_avg,rouge_2_score_avg,rouge_l_score_avg,chrf_score_avg

def calculate_shortanswer(min_row,max_row,answer_sheet,test_sheet):
    """
    计算单简答题题的分数。
    """
    # 按行遍历所有行
    num = 0
    bleu_score = [0, 0, 0, 0, 0]
    rouge_1_score = [0, 0, 0, 0, 0]
    rouge_2_score = [0, 0, 0, 0, 0]
    rouge_l_score = [0, 0, 0, 0, 0]
    chrf_score = [0, 0, 0, 0, 0]

    for index, row in enumerate(answer_sheet.iter_rows(min_row,max_row)):  # 假设第一行是标题行，从第二行开始读取
        num += 1
        # 读取每行中特定列的值
        reference = row[5].value.strip()  # 假设问题在第六列
        answer1 = test_sheet[index + min_row][25].value.strip()
        answer2 = test_sheet[index + min_row][26].value.strip()
        answer3 = test_sheet[index + min_row][27].value.strip()
        answer4 = test_sheet[index + min_row][28].value.strip()
        answer5 = test_sheet[index + min_row][29].value.strip()

        a1, a2, a3, a4, a5 = one_score(reference, answer1)
        b1, b2, b3, b4, b5 = one_score(reference, answer2)
        c1, c2, c3, c4, c5 = one_score(reference, answer3)
        d1, d2, d3, d4, d5 = one_score(reference, answer4)
        e1, e2, e3, e4, e5 = one_score(reference, answer5)

        bleu_score[0] += a1
        bleu_score[1] += b1
        bleu_score[2] += c1
        bleu_score[3] += d1
        bleu_score[4] += e1

        rouge_1_score[0] += a2
        rouge_1_score[1] += b2
        rouge_1_score[2] += c2
        rouge_1_score[3] += d2
        rouge_1_score[4] += e2

        rouge_2_score[0] += a3
        rouge_2_score[1] += b3
        rouge_2_score[2] += c3
        rouge_2_score[3] += d3
        rouge_2_score[4] += e3

        rouge_l_score[0] += a4
        rouge_l_score[1] += b4
        rouge_l_score[2] += c4
        rouge_l_score[3] += d4
        rouge_l_score[4] += e4

        chrf_score[0] += a5
        chrf_score[1] += b5
        chrf_score[2] += c5
        chrf_score[3] += d5
        chrf_score[4] += e5
        # print(reference)

    # print(num)
    # print(f'bleu_score_avg:{bleu_score[0] / num}  {bleu_score[1] / num}  {bleu_score[2] / num}  {bleu_score[3] / num}  {bleu_score[4] / num}')
    # print(f'rouge_1_score_avg:{rouge_1_score[0] / num}  {rouge_1_score[1] / num}  {rouge_1_score[2] / num}  {rouge_1_score[3] / num}  {rouge_1_score[4] / num}')
    # print(f'rouge_2_score_avg:{rouge_2_score[0] / num}  {rouge_2_score[1] / num}  {rouge_2_score[2] / num}  {rouge_2_score[3] / num}  {rouge_2_score[4] / num}')
    # print(f'rouge_l_score_avg:{rouge_l_score[0] / num}  {rouge_l_score[1] / num}  {rouge_l_score[2] / num}  {rouge_l_score[3] / num}  {rouge_l_score[4] / num}')
    # print(f'chrf_score_avg:{chrf_score[0] / num}  {chrf_score[1] / num}  {chrf_score[2] / num}  {chrf_score[3] / num}  {chrf_score[4] / num}')
    # print('简答题：')
    bleu_score_avg=(bleu_score[0]+bleu_score[1]+bleu_score[2]+bleu_score[3]+bleu_score[4]) /num/ 5
    rouge_1_score_avg=(rouge_1_score[0]+rouge_1_score[1]+rouge_1_score[2]+rouge_1_score[3]+rouge_1_score[4])/num/5
    rouge_2_score_avg=(rouge_2_score[0]+rouge_2_score[1]+rouge_2_score[2]+rouge_2_score[3]+rouge_2_score[4])/num/5
    rouge_l_score_avg=(rouge_l_score[0]+rouge_l_score[1]+rouge_l_score[2]+rouge_l_score[3]+rouge_l_score[4])/num/5
    chrf_score_avg=(chrf_score[0]+chrf_score[1]+chrf_score[2]+chrf_score[3]+chrf_score[4])/num/5
    # print(f'bleu_score_avg:{bleu_score_avg}')
    # print(f'rouge_1_score_avg:{rouge_1_score_avg}')
    # print(f'rouge_2_score_avg:{rouge_2_score_avg}')
    # print(f'rouge_l_score_avg:{rouge_l_score_avg}')
    # print(f'chrf_score_avg:{chrf_score_avg}')

    result_json["Shortanswer_bleu_score_avg"] = bleu_score_avg
    result_json["Shortanswer_rouge_1_score_avg"] = rouge_1_score_avg
    result_json["Shortanswer_rouge_2_score_avg"] = rouge_2_score_avg
    result_json["Shortanswer_rouge_l_score_avg"] = rouge_l_score_avg
    result_json["Shortanswer_chrf_score_avg"] = chrf_score_avg

    return bleu_score_avg,rouge_1_score_avg,rouge_2_score_avg,rouge_l_score_avg,chrf_score_avg

def calculate_connect_the_dots(min_row,max_row,answer_sheet,test_sheet):
    # answer = openpyxl.load_workbook(answer_sheet)
    # sheet1 = answer.active  # 假设数据在第一个工作表
    #
    # wb=openpyxl.load_workbook(test_sheet)
    # sheet2 = wb.active  # 假设数据在第一个工作表
    sheet1=answer_sheet
    sheet2=test_sheet

    acc_sum=[]
    top5_sum=[]
    # 按行遍历所有行
    for id in range(5):
        temp = []
        for index,row in enumerate(sheet1.iter_rows(min_row,max_row)):  # 假设第一行是标题行，从第二行开始读取
            # 使用ast.literal_eval()将字符串转换为Python列表
            try:
                list_data = ast.literal_eval(row[6].value)
            except (ValueError, SyntaxError):
                print("字符串格式不正确，无法转换为列表")

            model_data=ast.literal_eval(sheet2[index+min_row][30+id].value)
            # print(list_data)
            # print(model_data)

            for model_group in model_data:
                for true_group in list_data:
                    if model_group[0] in true_group[0]:
                        flag=True
                        # print(model_group)
                        # print(true_group)
                        for sub_index,model_answer in enumerate(model_group):
                            if not (model_answer in true_group[sub_index]):
                                flag=False
                                break
                        if flag==True:
                            temp.append(model_group)
                            top5_sum.append(model_group)
                            # print(true_group)
                            # print(f'acc:{model_group}')
                acc_sum.append(temp)
    # print(len(set(acc_sum[0]))/355)
    # print(len(set(acc_sum[1]))/355)
    # print(len(set(acc_sum[2]))/355)
    # print(len(set(acc_sum[3]))/355)
    # print(len(set(acc_sum[4]))/355)
    # print(f'连线_top1_avg{(len(set(acc_sum[0])) / 355+len(set(acc_sum[1])) / 355+len(set(acc_sum[2])) / 355+len(set(acc_sum[3])) / 355+len(set(acc_sum[4])) / 355)/5}')
    # print(f'连线_top5：{len(set(top5_sum))/355}')

    result_json["Connect_Top1"] = (len(set(acc_sum[0])) / 355+len(set(acc_sum[1])) / 355+len(set(acc_sum[2])) / 355+len(set(acc_sum[3])) / 355+len(set(acc_sum[4])) / 355)/5
    result_json["Connect_Top5"] = len(set(top5_sum))/355

    return (len(set(acc_sum[0])) / 355+len(set(acc_sum[1])) / 355+len(set(acc_sum[2])) / 355+len(set(acc_sum[3])) / 355+len(set(acc_sum[4])) / 355)/5


if __name__ == '__main__':
    # 加载权重文件
    weight = openpyxl.load_workbook(r'./weight.xlsx')
    weight_sheet = weight.active  # 假设数据在第一个工作表
    # 加载答案文件
    answer = openpyxl.load_workbook(r'./answer.xlsx')
    answer_sheet = answer.active  # 假设数据在第一个工作表
    # 加载测试文件
    # test = openpyxl.load_workbook(r'E:\题库\测试\结果\汇总答案/llama3-70b_测试文件.xlsx')
    test = openpyxl.load_workbook(r'./test.xlsx')
    test_sheet = test.active  # 假设数据在第一个工作表


    score_4=calculate_4choices(min_row=2,max_row=1162,answer_sheet=answer_sheet, test_sheet=test_sheet,weight_sheet=weight_sheet)
    score_5=calculate_5choices(min_row=2,max_row=403, answer_sheet=answer_sheet, test_sheet=test_sheet,weight_sheet=weight_sheet)
    score_judge=calculate_judgechoices(min_row=2,max_row=318, answer_sheet=answer_sheet, test_sheet=test_sheet,weight_sheet=weight_sheet)
    score_multi=calculate_multichoices(min_row=2, max_row=61, answer_sheet=answer_sheet, test_sheet=test_sheet)

    bleu_score_avg,rouge_1_score_avg,rouge_2_score_avg,rouge_l_score_avg,chrf_score_avg=calculate_terminology(min_row=2, max_row=184, answer_sheet=answer_sheet, test_sheet=test_sheet)
    score_terminology=Score_beita(bleu_score_avg,rouge_1_score_avg,rouge_2_score_avg,rouge_l_score_avg,chrf_score_avg)

    bleu_score_avg,rouge_1_score_avg,rouge_2_score_avg,rouge_l_score_avg,chrf_score_avg=calculate_shortanswer(min_row=2, max_row=232, answer_sheet=answer_sheet, test_sheet=test_sheet)
    score_shortanswer = Score_beita(bleu_score_avg, rouge_1_score_avg, rouge_2_score_avg, rouge_l_score_avg,chrf_score_avg)

    score_connect=calculate_connect_the_dots(min_row=2, max_row=79, answer_sheet=answer_sheet, test_sheet=test_sheet)
    DGscore(score_4, score_5, score_multi, score_judge, score_connect, score_terminology, score_shortanswer)

    restult.append(result_json)
    restult.append(result_com_json)
    print(restult)





