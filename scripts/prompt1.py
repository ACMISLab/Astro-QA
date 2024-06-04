#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/3/15 20:07
# @Author : 桐
# @QQ:1041264242
# 注意事项：

from http import HTTPStatus
import dashscope
import json
import openpyxl
model=['qwen1.5-32b-chat','qwen1.5-110b-chat']

def call_with_prompt(prompt,m):
    ###这里LLM的API调用接口和返回结果

if __name__ == '__main__':
    for m in model:
        # 加载xlsx文件
        wb = openpyxl.load_workbook(r'./天文学多选题.xlsx')
        sheet = wb.active  # 假设数据在第一个工作表
        # 假设results是一个二维列表，用于存储每次的回答结果
        results = []
        # 按行遍历所有行
        for row in sheet.iter_rows(min_row=2):  # 假设第一行是标题行，从第二行开始读取
            # 读取每行中特定列的值
            question = row[1].value.strip()  # 假设问题在第二列
            try:
                choice_A = row[2].value.strip()
            except:
                choice_A = row[2].value
            try:
                choice_B = row[3].value.strip()
            except:
                choice_B = row[3].value
            try:
                choice_C = row[4].value.strip()
            except:
                choice_C = row[4].value
            try:
                choice_D = row[5].value.strip()
            except:
                choice_D = row[5].value
            # try:
            #     choice_E = row[6].value.strip()
            # except:
            #     choice_E = row[6].value
            # print(question)
            # 对每个问题循环5次
            answers_for_question = []
            for id in range(5):
                # prompt_panduan = f'以下是天文学科目的判断题，请针对以下陈述进行判断，并严格按照以下格式回答：\n如果陈述正确，请回复："该陈述是正确的。"\n如果陈述错误，请回复："该陈述是错误的。"\n陈述内容："{question}"\n'
                prompt_duoxuan=f'请仔细阅读以下天文学科目的题目，每个问题可能有多个正确答案。请确保你的回答准确、完整，并附上简要的解释说明为什么选择这些答案。\n题目： "{question}"\n选项：\nA. {choice_A}\nB. {choice_B}\nC. {choice_C}\nD. {choice_D}\n请严格按照以下格式回答：\n答案是：在此处写下你的答案，例如(假设所有选项都正确):A、B、C、D\n解释说明：在此处写下你的选择的原因'
                # prompt_4xuan=f'请仔细阅读以下天文学科目的题目，并从给出的选项中选出正确的答案。每个问题只有一个正确答案。请确保你的回答准确，并附上简要的解释说明为什么选择这个答案。\n题目： "{question}"\n选项：\nA. {choice_A}\nB. {choice_B}\nC. {choice_C}\nD. {choice_D}\n请严格按照以下格式回答：\n答案是：在此处写下你的答案，例如(假设D选项正确):D\n解释说明：在此处写下你的选择的原因'
                # prompt_5xuan = f'请仔细阅读以下天文学科目的题目，并从给出的选项中选出正确的答案。每个问题只有一个正确答案。请确保你的回答准确，并附上简要的解释说明为什么选择这个答案。\n题目： "{question}"\n选项：\nA. {choice_A}\nB. {choice_B}\nC. {choice_C}\nD. {choice_D}\nE. {choice_E}\n请严格按照以下格式回答：\n答案是：在此处写下你的答案，例如(假设D选项正确):D\n解释说明：在此处写下你的选择的原因'

                # prompt_syjd=question
                if id ==0:
                    print(prompt_duoxuan)
                response=call_with_prompt(prompt_duoxuan,m)
                print(f'结果{id}：{response}')
                # 将答案添加到当前问题的答案列表中
                answers_for_question.append(response)
                # 将当前问题的所有答案添加到总结果列表中
            results.append(answers_for_question)

            # 将结果写入到第10到15列
        for i, answers in enumerate(results, start=2):  # start=2是因为我们从第二行开始写入数据
            for j, answer in enumerate(answers, start=10):  # start=10是因为我们要从第10列开始写入
                sheet.cell(row=i, column=j, value=answer)
                # 保存工作簿
        wb.save(r'result.xlsx')